"""Report service: generate, list, detail, export reconciliation reports (US8).

Aggregates bill, binding, and contribution data into multi-dimensional reports.
Excel export uses openpyxl.
"""

import io
import uuid
from datetime import datetime, timezone
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from ..core.exceptions import BadRequestException, NotFoundException
from ..models.bill import Bill, TransactionStatus
from ..models.binding import BindingStatus, Customer
from ..models.contribution import ContributionRecord, ContributionStatus
from ..models.distributor import Distributor
from ..models.organization import Organization
from ..models.report import Report


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
MAX_DATE_RANGE_DAYS = 366  # ~1 year


class ReportService:
    """Multi-dimensional reconciliation report service."""

    # ------------------------------------------------------------------
    # Validation
    # ------------------------------------------------------------------
    @staticmethod
    def _validate_date_range(start_date: str, end_date: str) -> tuple[datetime, datetime]:
        """Validate and parse date range. Raises BadRequestException on invalid input."""
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            end = datetime.strptime(end_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        except ValueError:
            raise BadRequestException(message="Invalid date format. Use YYYY-MM-DD.")

        if start > end:
            raise BadRequestException(message="Start date must be before end date.")

        if start > datetime.now(timezone.utc):
            raise BadRequestException(message="Date range cannot be entirely in the future.")

        if (end - start).days > MAX_DATE_RANGE_DAYS:
            raise BadRequestException(message=f"Date range cannot exceed {MAX_DATE_RANGE_DAYS} days.")

        return start, end

    # ------------------------------------------------------------------
    # Generate report
    # ------------------------------------------------------------------
    async def generate_report(
        self,
        db: AsyncSession,
        start_date: str,
        end_date: str,
        dimensions: list[str],
        user_display: Optional[str] = None,
    ) -> dict:
        """Generate a multi-dimensional reconciliation report.

        Aggregates data for each requested dimension and stores the report.
        """
        start, end = self._validate_date_range(start_date, end_date)

        report_id = uuid.uuid4().hex
        sections = {}

        for dim in dimensions:
            if dim == "binding":
                sections["binding"] = await self._build_binding_section(db, start, end)
            elif dim == "revenue":
                sections["revenue"] = await self._build_revenue_section(db, start, end)
            elif dim == "discount":
                sections["discount"] = await self._build_discount_section(db, start, end)
            elif dim == "allocation":
                sections["allocation"] = await self._build_allocation_section(db, start, end)

        now = datetime.now(timezone.utc)

        report = Report(
            id=report_id,
            start_date=start_date,
            end_date=end_date,
            dimensions=dimensions,
            sections=sections,
            generated_by=user_display,
            generated_at=now,
        )
        db.add(report)
        await db.flush()

        return {
            "reportId": report_id,
            "generatedAt": now,
            "dimensions": dimensions,
        }

    # ------------------------------------------------------------------
    # List reports
    # ------------------------------------------------------------------
    async def list_reports(self, db: AsyncSession) -> dict:
        """List all historical reports ordered by most recent first."""
        result = await db.execute(
            select(Report).order_by(Report.generated_at.desc()).limit(50)
        )
        reports = result.scalars().all()

        items = []
        for r in reports:
            items.append({
                "reportId": r.id,
                "dateRange": {"startDate": r.start_date, "endDate": r.end_date},
                "dimensions": r.dimensions or [],
                "generatedAt": r.generated_at.isoformat() if r.generated_at else None,
                "generatedBy": r.generated_by,
            })

        return {"items": items}

    # ------------------------------------------------------------------
    # Report detail
    # ------------------------------------------------------------------
    async def get_detail(self, db: AsyncSession, report_id: str) -> dict:
        """Get full report detail with all sections."""
        result = await db.execute(
            select(Report).where(Report.id == report_id)
        )
        report = result.scalars().first()
        if report is None:
            raise NotFoundException(message="Report not found")

        return {
            "reportId": report.id,
            "dateRange": {"startDate": report.start_date, "endDate": report.end_date},
            "dimensions": report.dimensions or [],
            "sections": report.sections or {},
            "generatedAt": report.generated_at.isoformat() if report.generated_at else None,
        }

    # ------------------------------------------------------------------
    # Export to Excel
    # ------------------------------------------------------------------
    async def export_excel(self, db: AsyncSession, report_id: str) -> bytes:
        """Export a report as an Excel (.xlsx) file.

        Returns the raw bytes of the Excel file.
        """
        detail = await self.get_detail(db, report_id)

        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # Style definitions
        header_font = Font(name="微软雅黑", bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Cover sheet
        ws_cover = wb.create_sheet("报表概览")
        ws_cover.merge_cells("A1:D1")
        ws_cover["A1"] = "北京儒泰分销管理系统 - 对账报表"
        ws_cover["A1"].font = Font(name="微软雅黑", bold=True, size=16)
        ws_cover["A1"].alignment = Alignment(horizontal="center")

        ws_cover["A3"] = "报表编号"
        ws_cover["B3"] = report_id
        ws_cover["A4"] = "日期范围"
        ws_cover["B4"] = f"{detail['dateRange']['startDate']} ~ {detail['dateRange']['endDate']}"
        ws_cover["A5"] = "维度"
        ws_cover["B5"] = ", ".join(detail["dimensions"])
        ws_cover["A6"] = "生成时间"
        ws_cover["B6"] = detail.get("generatedAt", "")
        for row in range(3, 7):
            ws_cover[f"A{row}"].font = Font(name="微软雅黑", bold=True)

        # Dimension sections
        dimension_sheet_config = {
            "binding": "绑定汇总",
            "revenue": "收入汇总",
            "discount": "优惠汇总",
            "allocation": "分配明细",
        }

        for dim_key, sheet_name in dimension_sheet_config.items():
            section = detail.get("sections", {}).get(dim_key)
            if not section:
                continue

            ws = wb.create_sheet(sheet_name)

            # Title
            ws.merge_cells("A1:F1")
            ws["A1"] = section.get("title", sheet_name)
            ws["A1"].font = Font(name="微软雅黑", bold=True, size=14)
            ws["A1"].alignment = Alignment(horizontal="center")

            # Summary row
            row = 3
            summary = section.get("summary", {})
            if summary:
                ws.cell(row=row, column=1, value="指标").font = header_font_white
                ws.cell(row=row, column=1).fill = header_fill
                ws.cell(row=row, column=2, value="数值").font = header_font_white
                ws.cell(row=row, column=2).fill = header_fill
                row += 1
                for k, v in summary.items():
                    ws.cell(row=row, column=1, value=k)
                    ws.cell(row=row, column=2, value=str(v))
                    row += 1

            # Details table
            details = section.get("details", [])
            if details:
                row += 1
                # Headers from first detail item keys
                if details:
                    headers = list(details[0].keys())
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col_idx, value=header)
                        cell.font = header_font_white
                        cell.fill = header_fill
                        cell.border = thin_border
                    row += 1

                    for detail_row in details:
                        for col_idx, header in enumerate(headers, 1):
                            val = detail_row.get(header, "")
                            cell = ws.cell(row=row, column=col_idx, value=str(val) if val is not None else "")
                            cell.border = thin_border
                        row += 1

            # Auto-width columns
            for col_idx in range(1, 10):
                ws.column_dimensions[get_column_letter(col_idx)].width = 20

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # ------------------------------------------------------------------
    # Section builders
    # ------------------------------------------------------------------
    async def _org_maps(
        self, db: AsyncSession, distributor_ids: set[int]
    ) -> tuple[dict[int, int], dict[int, str]]:
        """Return (distributor_id -> org_id, org_id -> org_name) maps."""
        org_map: dict[int, int] = {}
        org_name_map: dict[int, str] = {}
        if not distributor_ids:
            return org_map, org_name_map

        dists = (
            await db.execute(
                select(Distributor).where(Distributor.id.in_(distributor_ids))
            )
        ).scalars().all()
        org_ids = {d.org_id for d in dists}
        if org_ids:
            orgs = (
                await db.execute(
                    select(Organization).where(Organization.id.in_(org_ids))
                )
            ).scalars().all()
            org_name_map = {o.id: o.name for o in orgs}
        for d in dists:
            org_map[d.id] = d.org_id
        return org_map, org_name_map

    async def _org_level_map(self, db: AsyncSession, org_ids: set[int]) -> dict[int, int]:
        """Return org_id -> level for the given org ids."""
        if not org_ids:
            return {}
        orgs = (
            await db.execute(
                select(Organization).where(Organization.id.in_(org_ids))
            )
        ).scalars().all()
        return {o.id: o.level for o in orgs}

    async def _build_binding_section(
        self, db: AsyncSession, start: datetime, end: datetime
    ) -> dict:
        """Build binding summary section grouped by org (US6-AC6)."""
        result = await db.execute(
            select(Customer).where(
                Customer.bound_at >= start,
                Customer.bound_at <= end,
                Customer.binding_status == BindingStatus.BOUND,
            )
        )
        bindings = result.scalars().all()

        total_bindings = len(bindings)

        org_map, org_name_map = await self._org_maps(
            db, {c.distributor_id for c in bindings}
        )

        # Group by org
        by_org: dict[int | None, int] = {}
        for c in bindings:
            oid = org_map.get(c.distributor_id)
            by_org[oid] = by_org.get(oid, 0) + 1

        details = []
        for oid, count in sorted(by_org.items(), key=lambda x: -x[1]):
            name = org_name_map.get(oid) if oid is not None else None
            details.append({
                "组织": name or (f"Org {oid}" if oid is not None else "未分配组织"),
                "新绑定数": count,
            })

        return {
            "title": "绑定汇总",
            "summary": {"总新绑定数": total_bindings},
            "details": details,
        }

    async def _build_revenue_section(
        self, db: AsyncSession, start: datetime, end: datetime
    ) -> dict:
        """Build revenue summary section from bills."""
        result = await db.execute(
            select(Bill).where(
                Bill.transaction_time >= start,
                Bill.transaction_time <= end,
            )
        )
        bills = result.scalars().all()

        total_amount = sum(b.total_amount_cent for b in bills)
        total_paid = sum(b.paid_amount_cent for b in bills)
        total_refund = sum(b.refund_amount_cent for b in bills)
        transaction_count = len(bills)

        # Monthly breakdown
        by_month = {}
        for b in bills:
            m = b.transaction_time.strftime("%Y-%m")
            if m not in by_month:
                by_month[m] = {"count": 0, "totalAmount": 0, "paidAmount": 0, "refundAmount": 0}
            by_month[m]["count"] += 1
            by_month[m]["totalAmount"] += b.total_amount_cent
            by_month[m]["paidAmount"] += b.paid_amount_cent
            by_month[m]["refundAmount"] += b.refund_amount_cent

        details = []
        for m in sorted(by_month.keys()):
            d = by_month[m]
            details.append({
                "月份": m,
                "交易数": d["count"],
                "总金额(元)": f"{d['totalAmount'] / 100:.2f}",
                "实收(元)": f"{d['paidAmount'] / 100:.2f}",
                "退款(元)": f"{d['refundAmount'] / 100:.2f}",
            })

        return {
            "title": "收入汇总",
            "summary": {
                "总交易数": transaction_count,
                "总金额(元)": f"{total_amount / 100:.2f}",
                "实收总额(元)": f"{total_paid / 100:.2f}",
                "退款总额(元)": f"{total_refund / 100:.2f}",
            },
            "details": details,
        }

    async def _build_discount_section(
        self, db: AsyncSession, start: datetime, end: datetime
    ) -> dict:
        """Build discount summary section from bills."""
        result = await db.execute(
            select(Bill).where(
                Bill.transaction_time >= start,
                Bill.transaction_time <= end,
                Bill.discount_amount_cent > 0,
            )
        )
        bills = result.scalars().all()

        total_discount = sum(b.discount_amount_cent for b in bills)
        discounted_count = len(bills)

        details = []
        for b in bills[:100]:  # limit detail rows
            details.append({
                "交易号": b.transaction_id,
                "交易时间": b.transaction_time.isoformat() if b.transaction_time else "",
                "总金额(元)": f"{b.total_amount_cent / 100:.2f}",
                "优惠(元)": f"{b.discount_amount_cent / 100:.2f}",
                "实收(元)": f"{b.paid_amount_cent / 100:.2f}",
            })

        return {
            "title": "优惠汇总",
            "summary": {
                "优惠交易数": discounted_count,
                "优惠总额(元)": f"{total_discount / 100:.2f}",
            },
            "details": details,
        }

    async def _build_allocation_section(
        self, db: AsyncSession, start: datetime, end: datetime
    ) -> dict:
        """Build allocation section grouped by org dimension (US6-AC6)."""
        result = await db.execute(
            select(ContributionRecord).where(
                ContributionRecord.occurred_at >= start,
                ContributionRecord.occurred_at <= end,
            )
        )
        records = result.scalars().all()

        org_map, org_name_map = await self._org_maps(
            db, {r.distributor_id for r in records}
        )
        org_level_map = await self._org_level_map(db, set(org_name_map.keys()))

        # Group by org
        by_org: dict[int | None, dict] = {}
        for r in records:
            oid = org_map.get(r.distributor_id)
            agg = by_org.setdefault(oid, {"totalPoints": 0.0, "count": 0, "members": set()})
            try:
                agg["totalPoints"] += float(r.points)
            except (ValueError, TypeError):
                pass
            agg["count"] += 1
            agg["members"].add(r.distributor_id)

        details = []
        for oid, data in sorted(by_org.items(), key=lambda x: -x[1]["totalPoints"]):
            name = org_name_map.get(oid) if oid is not None else None
            level = org_level_map.get(oid) if oid is not None else None
            details.append({
                "组织": name or (f"Org {oid}" if oid is not None else "未分配组织"),
                "层级": f"L{level}" if level is not None else "N/A",
                "贡献值": f"{data['totalPoints']:.2f}",
                "记录数": data["count"],
                "分销员数": len(data["members"]),
            })

        return {
            "title": "分配明细",
            "summary": {
                "总记录数": len(records),
                "总组织数": len(by_org),
                "总分销员数": len({r.distributor_id for r in records}),
            },
            "details": details,
        }
